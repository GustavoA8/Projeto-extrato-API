import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from utils import resource_path



def looks_like_html(path, nbytes=512):
    try:
        with open(path, 'rb') as f:
            start = f.read(nbytes).lower()
        return b'<html' in start or b'<!doctype html' in start
    except Exception:
        return False

def read_with_fallback(path):
    # Try to read as a native Excel file first. If xlrd fails with BOF or
    # the content looks like HTML, try pandas.read_html as a fallback.
    try:
        if path.lower().endswith('.xlsx'):
            return pd.read_excel(path, header=None, engine='openpyxl')
        else:
            return pd.read_excel(path, header=None, engine='xlrd')
    except Exception as e:
        msg = str(e)
        # Common xlrd BOF/corrupt message contains 'Expected BOF' or 'Unsupported format'
        if 'expected bof' in msg.lower() or 'unsupported format' in msg.lower() or looks_like_html(path):
            print('Arquivo parece ser HTML ou está no formato "Web Page (HTML)". Tentando pd.read_html() como fallback...')
            try:
                tables = pd.read_html(path)
                if not tables:
                    raise RuntimeError('Nenhuma tabela encontrada pelo read_html')
                # If multiple tables found, concatenate them (you may choose a specific one)
                df = pd.concat(tables, ignore_index=True)
                return df
            except Exception as e2:
                print('Falha ao tentar ler como HTML:', e2)
                raise
        else:
            raise

def carregar_mapa_grupos(banco):
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    caminho = os.path.join(base_path, "grupo-aplic.xlsx")

    df_mapa = pd.read_excel(caminho, sheet_name=banco)

    df_mapa["Descrição"] = df_mapa["Descrição"].astype(str).str.upper().str.strip()

    return dict(zip(df_mapa["Descrição"], df_mapa["Grupo"]))

def separar_por_grupo(obj, banco):

    mapeamento = carregar_mapa_grupos(banco)

    aplicacoes_por_grupo = {}
    lancamentos_gerais = []

    for reg in obj:

        descricao = str(reg["lancamento"]).upper().strip()
        grupo_encontrado = None

        for desc_mapa, grupo in mapeamento.items():
            if desc_mapa in descricao:
                grupo_encontrado = grupo
                break

        if grupo_encontrado:
            if grupo_encontrado not in aplicacoes_por_grupo:
                aplicacoes_por_grupo[grupo_encontrado] = []

            aplicacoes_por_grupo[grupo_encontrado].append(reg)
        elif 'tar' not in descricao.lower():
            lancamentos_gerais.append(reg)

    return aplicacoes_por_grupo, lancamentos_gerais

def localizar_texto_df(df, texto):
    encontrados = []

    for i, linha in df.iterrows():
        for j, valor in linha.items():
            if pd.notna(valor) and texto.lower() in str(valor).lower():
                encontrados.append((i, j, valor))

    return encontrados

def pegar_valor_cabecalho(df, linha, col_inicial):
    # percorre as colunas depois da encontrada
    for col in range(col_inicial + 1, len(df.columns)):
        valor = df.iloc[linha, col]

        if pd.notna(valor) and str(valor).strip() != "":
            return valor

    return None

def padronizar_extrato_itau(df):

    # 1 - achar linha do cabeçalho
    h = None

    for i, linha in df.iterrows():

        textos = [str(v).lower() for v in linha if pd.notna(v)]

        if ("data" in " ".join(textos) and
            "saldo" in " ".join(textos)):
            h = i
            break

    if h is None:
        print("Não achei linha de cabeçalho")
        return None


    # 2 - mapear colunas
    header = df.iloc[h]

    col_data = None
    col_lanc = None
    col_valor = None
    col_saldo = None

    for j, v in header.items():
        if pd.isna(v):
            continue

        v = str(v).lower()

        if "data" in v:
            col_data = j

        elif "lan" in v:
            col_lanc = j

        elif "valor" in v:
            col_valor = j

        elif "saldo" in v:
            col_saldo = j


    # 3 - validação
    if None in (col_data, col_lanc, col_valor, col_saldo):
        print("Mapeamento falhou:",
              col_data, col_lanc, col_valor, col_saldo)
        return None


    # 4 - cortar até último saldo
    linha_fim, _ = ultima_celula_saldo(df, col_saldo)

    dados = df.iloc[h+1 : linha_fim+1]


    # 5 - montar dataframe final
    novo = pd.DataFrame({
        "Data": dados.iloc[:, col_data],
        "Lançamento": dados.iloc[:, col_lanc],
        "Valor": dados.iloc[:, col_valor],
        "Saldo": dados.iloc[:, col_saldo],
    })

    return novo.dropna(how='all')

def padronizar_extrato_santander(df):

    # 1 - achar linha do cabeçalho
    h = None

    for i, linha in df.iterrows():

        textos = [str(v).lower() for v in linha if pd.notna(v)]

        if ("data" in " ".join(textos) and
            "saldo" in " ".join(textos)):
            h = i
            break

    if h is None:
        print("Não achei linha de cabeçalho")
        return None


    # 2 - mapear colunas
    header = df.iloc[h]

    col_data = None
    col_lanc = None
    col_valor = None
    col_saldo = None

    for j, v in header.items():
        if pd.isna(v):
            continue

        v = str(v).lower()

        if "data" in v:
            col_data = j

        elif "hist" in v:
            col_lanc = j

        elif "valor" in v:
            col_valor = j

        elif "saldo" in v:
            col_saldo = j


    # 3 - validação
    if None in (col_data, col_lanc, col_valor, col_saldo):
        print("Mapeamento falhou:",
              col_data, col_lanc, col_valor, col_saldo)
        return None


    # 4 - cortar até último saldo
    linha_fim, _ = ultima_celula_saldo(df, col_saldo)

    dados = df.iloc[h+1 : linha_fim+1]


    # 5 - montar dataframe final
    novo = pd.DataFrame({
        "Data": dados.iloc[:, col_data],
        "Lançamento": dados.iloc[:, col_lanc],
        "Valor": dados.iloc[:, col_valor],
        "Saldo": dados.iloc[:, col_saldo],
    })

    return novo.dropna(how='all')

def padronizar_extrato_bradesco(df):

    # 1 - achar linha do cabeçalho
    h = None

    for i, linha in df.iterrows():
        textos = [str(v).lower() for v in linha if pd.notna(v)]

        if "data" in " ".join(textos) and "saldo" in " ".join(textos):
            h = i
            break

    if h is None:
        print("Não achei linha de cabeçalho Bradesco")
        return None


    header = df.iloc[h]


    # 2 - mapear colunas
    col_data = None
    col_lanc = None
    col_credito = None
    col_debito = None
    col_saldo = None

    for j, v in header.items():
        if pd.isna(v):
            continue

        v = str(v).lower()

        if "data" in v:
            col_data = j

        elif "lan" in v or "hist" in v:
            col_lanc = j

        elif "crédito" in v or "credito" in v:
            col_credito = j

        elif "débito" in v or "debito" in v:
            col_debito = j

        elif "saldo" in v:
            col_saldo = j


    if None in (col_data, col_lanc, col_credito, col_debito, col_saldo):
        print("Mapeamento falhou Bradesco:",
              col_data, col_lanc, col_credito, col_debito, col_saldo)
        return None


    # 3 - achar linha TOTAL
    linha_fim = None

    for i in range(h, len(df)):
        linha = df.iloc[i]
        textos = [str(v).lower() for v in linha if pd.notna(v)]

        if any("total" in t for t in textos):
            linha_fim = i
            break

    if linha_fim is None:
        print("Não achei linha TOTAL do Bradesco")
        return None


    dados = df.iloc[h+1 : linha_fim].copy()


    # 4 - montar valor único com sinal
    valores = []

    for _, row in dados.iterrows():

        cred = converter_valor_br(row.iloc[col_credito])
        deb  = -(converter_valor_br(row.iloc[col_debito]))

        valor = cred - deb   # débito vira negativo

        valores.append(valor)


    # 5 - montar dataframe final
    novo = pd.DataFrame({
        "Data": dados.iloc[:, col_data],
        "Lançamento": dados.iloc[:, col_lanc],
        "Valor": valores,
        "Saldo": dados.iloc[:, col_saldo].apply(converter_valor_br),
    })


    return novo.dropna(how='all')

def converter_valor_br(valor):
    if pd.isna(valor):
        return 0.0

    v = str(valor).strip()

    # remove pontos de milhar e troca vírgula por ponto
    v = v.replace(".", "").replace(",", ".")

    try:
        return float(v)
    except:
        return 0.0

def extrato_para_obj(df):
    extrato = []

    for i, linha in df.iterrows():
        registro = {
            "data": linha["Data"],
            "lancamento": linha["Lançamento"],
            "valor": linha["Valor"],
            "saldo": linha["Saldo"],
        }
        extrato.append(registro)
            

    return extrato

def ultima_celula_saldo(df, coluna):
    for i in range(len(df) - 1, -1, -1):
        valor = df.iloc[i, coluna]

        if pd.notna(valor) and str(valor).strip() != "":
            return i, valor

    return None, None

def ultima_cedula_saldo_bradesco(df, coluna):
    df = df.fillna("").astype(str)

    for i, row in df.iterrows():
        # procura a linha que contém a palavra TOTAL
        if row.str.contains("total", case=False).any():
            valor = df.iloc[i, coluna]

            # tratamento de formato brasileiro
            valor_tratado = str(valor).replace(".", "").replace(",", ".")

            try:
                return i, float(valor_tratado)
            except:
                return i, valor

    return None, None
    df = df.fillna("").astype(str)

    for i, row in df.iterrows():
        if row.str.contains("total", case=False).any():
            valor = df.iloc[i, coluna_saldo]

            # converte para número se estiver com vírgula
            valor = str(valor).replace(".", "").replace(",", ".")

            try:
                return float(valor)
            except:
                return valor

    return None

def separador_lancamentos_tar(objeto_extrato):
    saldos = []
    for registro in objeto_extrato:
        if 'tar' in registro['lancamento'].lower() or 'custas' in registro['lancamento'].lower():
            saldos.append(registro)
    return saldos

def separador_lancamentos_aplic(objeto_extrato):
    aplic = []
    for registro in objeto_extrato:
        if 'aplic' in registro['lancamento'].lower() :
            aplic.append(registro)
    return aplic

def separador_lancamentos_contamax(objeto_extrato):
    contamax = []
    for registro in objeto_extrato:
        if 'contamax' in registro['lancamento'].lower():
            contamax.append(registro)
    return contamax

def separador_lancamentos_creditos(objeto_extrato):
    creditos = []
    for registro in objeto_extrato:
        if float(registro['valor']) > 0 and registro['lancamento'].lower() != 'aplic' and "cdb" not in registro['lancamento'].lower() and "aplic" not in registro['lancamento'].lower():
            creditos.append(registro)
    return creditos

def separador_lancamentos_credito(objeto_extrato):
    credito = []
    for registro in objeto_extrato:
        if 'cr cob' in registro['lancamento'].lower():
            credito.append(registro)
    return credito

def separar_lancamentos_geral(objeto_extrato):
    geral = []
    for registro in objeto_extrato:
        if 'tar' not in registro['lancamento'].lower() and 'contamax' not in registro['lancamento'].lower():
            geral.append(registro)
    return geral

def separador_lancamentos_debitos(objeto_extrato):
    debitos = []
    for registro in objeto_extrato:
        if float(registro['valor']) < 0 and 'tar' not in registro['lancamento'].lower() and 'contamax' not in registro['lancamento'].lower():
            debitos.append(registro)
    return debitos

def main_itau(df):
    print("---------------------------------Sheets SALDOS----------------------------------------")
    nome_loc = localizar_texto_df(df, 'Nome:')
    

    print(nome_loc)
    
    for linha, coluna, valor in nome_loc:
        nomeCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Nome encontrado: {valor}, Valor após: {nomeCond}")

    print(nomeCond)

    agenciaConta_loc = localizar_texto_df(df, 'Agência/Conta:')
    if not agenciaConta_loc:
        agencia_loc = localizar_texto_df(df, 'Agência:')
        conta_loc = localizar_texto_df(df, 'Conta:')
        for linha, coluna, valor in agencia_loc:
            agenciaCond = pegar_valor_cabecalho(df, linha, coluna)
            print(f"Agência encontrado: {valor}, Valor após: {agenciaCond}")
        for linha, coluna, valor in conta_loc:
            contaCond = pegar_valor_cabecalho(df, linha, coluna)
            print(f"Conta encontrada: {valor}, Valor após: {contaCond}")
        agenciaContaCond = f"{agenciaCond}/{contaCond}"
        print(f"Agência/Conta combinado: {agenciaContaCond}")
    else:
        print(agenciaConta_loc)
    for linha, coluna, valor in agenciaConta_loc:
        agenciaContaCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Agência/Conta encontrado: {valor}, Valor após: {agenciaContaCond}")


    
    



    data_loc = localizar_texto_df(df, 'Periodo:')
    if not data_loc:
        data_loc = localizar_texto_df(df, 'Atualização:')


    print(data_loc)
    for linha, coluna, valor in data_loc:
        dataCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Data encontrado: {valor}, Valor após: {dataCond}")

    saldoAnt_loc = localizar_texto_df(df, 'Saldo Anterior')
    print(saldoAnt_loc)

    for linha, coluna, valor in saldoAnt_loc:
        saldoAntCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Saldo Anterior encontrado: {valor}, Valor após: {saldoAntCond}")

    sdo_inicial = saldoAntCond
    sdo_final = 0

    colunaSaldo = localizar_texto_df(df, 'Saldo (R$)')
    print(colunaSaldo)
    if colunaSaldo:
        
        coluna_index = colunaSaldo[0][1]
        linha_index, sdo_final = ultima_celula_saldo(df, coluna_index)
        print(f"Último saldo na coluna 'Saldo (R$)' encontrado na linha {linha_index}: {sdo_final}")
    else:
        print("Coluna 'Saldo (R$)' não encontrada.")
    print(f"\nResumo:\nNome: {nomeCond}\nAgência/Conta: {agenciaContaCond}\nData: {dataCond[:10]}\nSaldo Inicial: {sdo_inicial}\nSaldo Anterior: {sdo_inicial}\nSaldo Final: {sdo_final}")
    print("---------------------------------Sheets Tarifas----------------------------------------")
    print("Extrato:")
    extrato_df = padronizar_extrato_itau(df)
    obj = extrato_para_obj(extrato_df)
    tarifa = separador_lancamentos_tar(obj)
    total = 0
    for registro in tarifa:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total += float(registro['valor'])

    print("Qtde: ", len(tarifa), " Total Tarifas/Custas: ", total.__round__(2))

    aplicacao = separador_lancamentos_aplic(obj)
    print("---------------------------------Sheets Aplicações----------------------------------------")
    total_aplic = 0
    for registro in aplicacao:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_aplic += float(registro['valor'])
    print("Qtde: ", len(aplicacao), " Total Aplicações: ", total_aplic.__round__(2))

    creditos = separador_lancamentos_creditos(obj)
    print("---------------------------------Sheets Créditos----------------------------------------")
    total_creditos = 0
    for registro in creditos:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_creditos += float(registro['valor'])
    print("Qtde: ", len(creditos), " Total Créditos: ", total_creditos.__round__(2))
    
    debitos = separador_lancamentos_debitos(obj)
    print("---------------------------------Sheets Débitos----------------------------------------")
    total_debitos = 0
    for registro in debitos:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_debitos += float(registro['valor'])
    print("Qtde: ", len(debitos), " Total Débitos: ", total_debitos.__round__(2))
    geral = separar_lancamentos_geral(obj)
    
    gerar_excel(
    nome=nomeCond,
    agencia_conta=agenciaContaCond,
    

    saldo_inicial=sdo_inicial,
    saldo_final=sdo_final,

    tarifas=tarifa,
    aplicacoes=aplicacao,
    creditos=creditos,
    debitos=debitos,
    geral=geral,
    arquivo_origem=df.attrs.get("arquivo_origem"),
    obj = obj,
    periodo=dataCond
    )

def main_santander(df, condominio):
    print("---------------------------------Sheets SALDOS----------------------------------------")
    agencia_loc = localizar_texto_df(df, 'Agencia')
    agenciaCond = df.loc[agencia_loc[0][0] , agencia_loc[0][1]+ 1]

    conta_loc = localizar_texto_df(df, 'Conta')
    contaCond = df.loc[conta_loc[0][0] , conta_loc[0][1]+ 1]


    agenciaContaCond = f"{agenciaCond.astype(int)}/{contaCond}"
    print(f"Agência/Conta combinado: {agenciaContaCond}")


    
    



    saldoAnt_loc = localizar_texto_df(df, 'Saldo Anterior')
    print(saldoAnt_loc)

    for linha, coluna, valor in saldoAnt_loc:
        saldoAntCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Saldo Anterior encontrado: {valor}, Valor após: {saldoAntCond}")

    sdo_inicial = saldoAntCond
    sdo_final = 0

    colunaSaldo = localizar_texto_df(df, 'Saldo (R$)')
    print(colunaSaldo)
    if colunaSaldo:
        
        coluna_index = colunaSaldo[0][1]
        linha_index, sdo_final = ultima_celula_saldo(df, coluna_index)
        print(f"Último saldo na coluna 'Saldo (R$)' encontrado na linha {linha_index}: {sdo_final}")
    else:
        print("Coluna 'Saldo (R$)' não encontrada.")
    print(f"\nResumo:\nAgência/Conta: {agenciaContaCond}\nSaldo Inicial: {sdo_inicial}\nSaldo Anterior: {sdo_inicial}\nSaldo Final: {sdo_final}")
    print("---------------------------------Sheets Tarifas----------------------------------------")
    print("Extrato:")
    extrato_df = padronizar_extrato_santander(df)
    obj = extrato_para_obj(extrato_df)
    tarifa = separador_lancamentos_tar(obj)
    total = 0
    for registro in tarifa:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total += float(registro['valor'])

    print("Qtde: ", len(tarifa), " Total Tarifas/Custas: ", total.__round__(2))

    
    print("---------------------------------Sheets CONTAMAX----------------------------------------")
    aplicacao = separador_lancamentos_aplic(obj)
    lista_aplicacao = []
    lista_resgate = []
    total_aplic = 0
    total_resgate = 0
   
 
    print("---------------------------------Sheets credito----------------------------------------")
    credito = separador_lancamentos_credito(obj)
    total_credito = 0
    for registro in credito:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_credito += float(registro['valor'])
    print("Qtde: ", len(credito), " Total credito]: ", total_credito.__round__(2))
    
    debitos = separador_lancamentos_debitos(obj)
    print("---------------------------------Sheets Débitos----------------------------------------")
    total_debitos = 0
    for registro in debitos:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_debitos += float(registro['valor'])
    print("Qtde: ", len(debitos), " Total Débitos: ", total_debitos.__round__(2))
    geral = separar_lancamentos_geral(obj) 

    gerar_excel(
    nome=condominio,
    banco="Santander",
    agencia_conta=agenciaContaCond,


    saldo_inicial=sdo_inicial,
    saldo_final=sdo_final,

    tarifas=tarifa,

    aplicacoes=aplicacao,

    credito=credito,

    debitos=debitos,
    geral=geral,

    arquivo_origem=df.attrs.get("arquivo_origem"),
    obj = obj
)

def main_bradesco(df, condominio):
    
    print("---------------------------------Sheets SALDOS----------------------------------------")
    print(df)
    agenciaConta_loc = localizar_texto_df(df, 'Agência/Conta:')
    if not agenciaConta_loc:
        agencia_loc = localizar_texto_df(df, 'Agência:')
        conta_loc = localizar_texto_df(df, 'Conta:')
        for linha, coluna, valor in agencia_loc:
            pos_ag = valor.find("Agência:") + len("Agência:")
            agenciaCond = valor[pos_ag:pos_ag + 6].strip()

            print(f"Agência encontrado: {valor}, Valor após: {agenciaCond}")
            

            
        for linha, coluna, valor in conta_loc:
            pos_ct = valor.find("Conta:") + len("Conta:")
            contaCond = valor[pos_ct:pos_ct+10].strip()
            print(f"Conta encontrada: {valor}, Valor após: {contaCond}")
        agenciaContaCond = f"{agenciaCond}/{contaCond}"
        print(f"Agência/Conta combinado: {agenciaContaCond}")
    else:
        print(agenciaConta_loc)
    for linha, coluna, valor in agenciaConta_loc:
        agenciaContaCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Agência/Conta encontrado: {valor}, Valor após: {agenciaContaCond}")

    saldoAnt_loc = localizar_texto_df(df, 'Saldo Anterior')
    print(saldoAnt_loc)

    for linha, coluna, valor in saldoAnt_loc:
        saldoAntCond = pegar_valor_cabecalho(df, linha, coluna)
        print(f"Saldo Anterior encontrado: {valor}, Valor após: {saldoAntCond}")
        break

    sdo_inicial = saldoAntCond
    sdo_final = 0

    colunaSaldo = localizar_texto_df(df, 'Saldo (R$)')
    print(colunaSaldo)
    if colunaSaldo:
        
        coluna_index = colunaSaldo[0][1]
        linha_index, sdo_final = ultima_cedula_saldo_bradesco(df, coluna_index)
        print(f"Último saldo na coluna 'Saldo (R$)' encontrado na linha {linha_index}: {sdo_final}")
    else:
        print("Coluna 'Saldo (R$)' não encontrada.")
    print(f"\nResumo:\nAgência/Conta: {agenciaContaCond}\nSaldo Inicial: {sdo_inicial}\nSaldo Anterior: {sdo_inicial}\nSaldo Final: {sdo_final}")
    print("---------------------------------Sheets Tarifas----------------------------------------")
    print("Extrato:")
    extrato_df = padronizar_extrato_bradesco(df)
    obj = extrato_para_obj(extrato_df)
    tarifa = separador_lancamentos_tar(obj)
    total = 0
    for registro in tarifa:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total += float(registro['valor'])

    print("Qtde: ", len(tarifa), " Total Tarifas/Custas: ", total.__round__(2))

    aplicacao = separador_lancamentos_aplic(obj)
    print("---------------------------------Sheets Aplicações----------------------------------------")
    total_aplic = 0
    for registro in aplicacao:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_aplic += float(registro['valor'])
    print("Qtde: ", len(aplicacao), " Total Aplicações: ", total_aplic.__round__(2))

    creditos = separador_lancamentos_creditos(obj)
    print("---------------------------------Sheets Créditos----------------------------------------")
    total_creditos = 0
    for registro in creditos:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_creditos += float(registro['valor'])
    print("Qtde: ", len(creditos), " Total Créditos: ", total_creditos.__round__(2))
    
    debitos = separador_lancamentos_debitos(obj)
    geral = separar_lancamentos_geral(obj)   
      
    print("---------------------------------Sheets Débitos----------------------------------------")
    total_debitos = 0
    for registro in debitos:
        print("Data: ", registro['data'], " Lançamento: ", registro['lancamento'], " Valor: ", registro['valor'], " Saldo: ", registro['saldo'])
        total_debitos += float(registro['valor'])
    print("Qtde: ", len(debitos), " Total Débitos: ", total_debitos.__round__(2))
   
    gerar_excel(
    nome=condominio,
    banco="Bradesco",
    agencia_conta=agenciaContaCond,

    saldo_inicial=sdo_inicial,
    saldo_final=sdo_final,
    tarifas=tarifa,
    aplicacoes=aplicacao,
    creditos=creditos,
    debitos=debitos,
    geral=geral,
    arquivo_origem=df.attrs.get("arquivo_origem"),
    obj = obj
    )

def gerar_excel(
    nome,
    banco,
    agencia_conta,
    saldo_inicial,
    saldo_final,
    tarifas,
    aplicacoes,
    creditos,
    debitos,
    geral,
    arquivo_origem,
    obj,
    periodo=None
):
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    from utils import resource_path

    wb = Workbook()

    base = os.path.basename(arquivo_origem)
    nome_sem_ext, ext = os.path.splitext(base)
    pasta = os.path.dirname(arquivo_origem)
    caminho = os.path.join(pasta, f"{nome_sem_ext}_extrato_conciliacao.xlsx")

    # ========== ESTILOS ==========
    font_titulo = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    font_cabecalho_tabela = Font(name='Arial', bold=True, size=10)
    font_normal = Font(name='Arial', size=10)
    font_total = Font(name='Arial', bold=True, size=10)
    font_negativo = Font(name='Arial', size=10, color="FF0000")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    borda = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)
    
    fill_cabecalho = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    fill_total = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    border_thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    borda_esquerda = Border(
        left=Side(style="thin", color="000000"),
        right=None,
        top=None,
        bottom=None
    )

    def configurar_impressao(ws, ultima_linha):
        # 📄 Papel A4
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # 📄 Orientação
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        # se quiser retrato: ws.ORIENTATION_PORTRAIT

        # 📄 Ajustar para caber em 1 página (largura)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False  # não força altura

        # 📄 Centralizar
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False

        # 📄 Margens
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # 📄 Área de impressão (ajusta automático)
        ws.print_area = f'B1:E{ultima_linha}'

        # 📄 Repetir cabeçalho da tabela (linha 12 no seu caso)
        ws.print_title_rows = '12:12'

    def montar_aba(nome_aba, lista, titulo_secao):
        creditos_lista = [x for x in lista if float(x["valor"]) > 0]
        debitos_lista = [x for x in lista if float(x["valor"]) < 0]
        ws = wb.create_sheet(nome_aba)
        ws.sheet_view.showGridLines = False
        if banco == "Itau":
            img = Image(resource_path("img/itau_header.png"))
        elif banco == "Santander":
            img = Image(resource_path("img/santander_header.png"))
        else:
            img = Image(resource_path("img/logo_bradesco.png"))
            img.width = 120   # largura
            img.height = 80   # altura
        ws.add_image(img, 'B1')
        
        # Logo/Cabeçalho
        ws.merge_cells('B2:C2')
        ws['B2'] = banco
        ws['B2'].font = Font(name='Arial', bold=True, size=14, color="0066CC")
        ws['B2'].alignment = align_left
        
        # Informações do cliente
        ws['B5'] = 'Nome:'
        ws['B5'].font = font_titulo
        ws['B5'].alignment = align_center
        ws['B5'].fill = PatternFill(fill_type='solid', start_color='1f497d')
        ws['C5'] = nome
        ws['C5'].font = font_normal
        
        ws['D5'] = 'Agência/Conta:'
        ws['D5'].font = font_titulo
        ws['D5'].alignment = align_center
        ws['D5'].fill = PatternFill(fill_type='solid', start_color='1f497d')
        ws['D6'] = agencia_conta
        ws['D6'].font = font_normal
        ws['D6'].border = borda_simples
        
        ws['B6'] = 'Data:'
        ws['B6'].font = font_titulo
        ws['B6'].alignment = align_center
        ws['B6'].fill = PatternFill(fill_type='solid', start_color='1f497d')
        ws['B6'].border = borda_simples
        ws['C6'] = periodo
        ws['C6'].font = font_normal
        ws['C6'].border = borda_simples

        for linha in range(5, 7):          # 5 até 6
           for col in ["B", "C"]:         # colunas B e C
              ws[f"{col}{linha}"].border = borda_simples
        
        # Título da seção
        ws['B9'] = titulo_secao
        ws['B9'].font = Font(name='Arial', bold=True, size=11)
        
        # Cabeçalho da tabela
        linha = 12

     # =============================
    # Separação de créditos/debitos
    # =============================
        if nome_aba == "Geral":

         creditos_lista = [x for x in lista if float(x["valor"]) > 0]
         debitos_lista = [x for x in lista if float(x["valor"]) < 0]

         listas = [
             ("CRÉDITOS", creditos_lista),
            ("DÉBITOS", debitos_lista)
         ]

        else:
         listas = [(titulo_secao, lista)]



        for titulo, dados in listas:
          tem_credito = any(float(x["valor"]) > 0 for x in dados)
          tem_debito = any(float(x["valor"]) < 0 for x in dados)
          
          if nome_aba == "Geral":
              ws[f"B{linha-2}"] = titulo
          
              ws[f"B{linha-2}"].font = Font(name='Arial', bold=True, size=11)

          if nome_aba == "Tarifas":
              headers = ["Data", "Lançamento", "Debito (R$)"]
              colunas = ["B", "C", "D"]
          else:
              if tem_credito and tem_debito:
                  headers = ["Data", "Lançamento", "Credito (R$)", "Debito (R$)"]
                  colunas = ["B", "C", "D", "E"]
              elif tem_credito:
                  headers = ["Data", "Lançamento", "Credito (R$)"]
                  colunas = ["B", "C", "D"]
              else:
                  headers = ["Data", "Lançamento", "Debito (R$)"]
                  colunas = ["B", "C", "D"]

          

          for col, texto in zip(colunas, headers):

              cell = ws[f"{col}{linha}"]
              cell.value = texto
              cell.font = font_cabecalho_tabela
              cell.alignment = align_center
              cell.fill = fill_cabecalho
              cell.border = border_thin

          linha += 1

          if not dados:

              ws[f"B{linha}"] = "Sem registros para esta categoria"
              ws[f"B{linha}"].font = Font(name='Arial', italic=True, size=10)
              linha += 2
              continue

          linha_inicio_dados = linha

          for reg in dados:

              valor = float(reg["valor"])
  
              ws[f"B{linha}"] = reg["data"]
              ws[f"B{linha}"].font = font_normal
              ws[f"B{linha}"].alignment = align_center
              ws[f"B{linha}"].border = border_thin

              ws[f"C{linha}"] = reg["lancamento"]
              ws[f"C{linha}"].font = font_normal
              ws[f"C{linha}"].alignment = align_left
              ws[f"C{linha}"].border = border_thin

              if nome_aba == "Tarifas":

                  ws[f"D{linha}"] = valor
                  ws[f"D{linha}"].font = font_negativo
                  
  
              else:
                  if tem_credito and tem_debito:
                      if valor > 0:
                        ws[f"E{linha}"] = valor
                        ws[f"E{linha}"].font = font_normal
                      else:
                        ws[f"E{linha}"] = valor
                        ws[f"E{linha}"].font = font_negativo
                  elif tem_credito:
                    ws[f"D{linha}"] = valor
                    ws[f"D{linha}"].font = font_normal
                  else:
                    ws[f"D{linha}"] = valor
                    ws[f"D{linha}"].font = font_negativo

              ws[f"D{linha}"].alignment = align_right
              ws[f"D{linha}"].number_format = '#,##0.00'
              ws[f"D{linha}"].border = border_thin

              if "E" in colunas:
                  ws[f"E{linha}"].alignment = align_right
                  ws[f"E{linha}"].number_format = '#,##0.00'
                  ws[f"E{linha}"].border = border_thin

              linha += 1

        # TOTAL
          ws.merge_cells(f'B{linha}:C{linha}')

          ws[f"B{linha}"] = "TOTAL"
          ws[f"B{linha}"].font = font_total
          ws[f"B{linha}"].alignment = align_center
          ws[f"B{linha}"].fill = fill_total
          ws[f"B{linha}"].border = border_thin

          ws[f"C{linha}"].border = border_thin

          ws[f"D{linha}"] = f'=SUM(D{linha_inicio_dados}:D{linha-1})'
          ws[f"D{linha}"].font = font_total
          ws[f"D{linha}"].alignment = align_right
          ws[f"D{linha}"].number_format = '#,##0.00'
          ws[f"D{linha}"].fill = fill_total
          ws[f"D{linha}"].border = border_thin

          if nome_aba != "Tarifas" and nome_aba != "Geral":
              ws[f"E{linha}"] = f'=SUM(E{linha_inicio_dados}:E{linha-1})'
              ws[f"E{linha}"].font = font_total
              ws[f"E{linha}"].alignment = align_right
              ws[f"E{linha}"].number_format = '#,##0.00'
              ws[f"E{linha}"].fill = fill_total
              ws[f"E{linha}"].border = border_thin

          linha += 3
        # ==========================
        # SALDOS (somente para aba GERAL)
        # ==========================
        if nome_aba == "Geral":
        
            ws[f'B{linha}'] = 'Saldo Inicial:'
            ws[f'B{linha}'].font = font_normal
            ws[f"B{linha}"].border = border_thin
        
            ws[f'C{linha}'] = saldo_inicial
            ws[f'C{linha}'].font = font_normal
            ws[f'C{linha}'].number_format = '#,##0.00'
            ws[f"C{linha}"].border = border_thin

            linha += 1

            ws[f'B{linha}'] = 'Saldo Final:'
            ws[f'B{linha}'].font = font_normal
            ws[f"B{linha}"].border = border_thin

            ws[f'C{linha}'] = saldo_final
            ws[f'C{linha}'].font = font_normal
            ws[f'C{linha}'].number_format = '#,##0.00'
            ws[f"C{linha}"].border = border_thin

            linha += 1

        

        
        # Ajustar larguras
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        if nome_aba == "Tarifas":
            ws.column_dimensions["E"].width = 8

        #configuração de impressão

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False

        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        configurar_impressao(ws, linha)

    def montar_aba_grupos(nome_aba, grupos_dict, titulo_secao, mapa_saldos_resumo=None):

         ws = wb.create_sheet(nome_aba)
         ws.sheet_view.showGridLines = False
         if banco == "Itau":
             img = Image(resource_path("img/itau_header.png"))
         elif banco == "Santander":
             img = Image(resource_path("img/santander_header.png"))
         else:
             img = Image(resource_path("img/logo_bradesco.png"))
         ws.add_image(img, 'B1')

         # Cabeçalho Itaú
         ws.merge_cells('B2:C2')
         ws['B2'] = banco
         ws['B2'].font = Font(name='Arial', bold=True, size=14, color="0066CC")
         ws['B2'].alignment = align_left

         # Informações
         ws['B5'] = 'Nome:'
         ws['B5'].font = font_titulo
         ws['B5'].alignment = align_center
         ws['B5'].fill = PatternFill(fill_type='solid', start_color='1f497d')
         ws['C5'] = nome
         ws['C5'].font = font_normal

         ws['D5'] = 'Agência/Conta:'
         ws['D5'].font = font_titulo
         ws['D5'].alignment = align_center
         ws['D5'].fill = PatternFill(fill_type='solid', start_color='1f497d')
         ws['E5'] = agencia_conta
         ws['E5'].font = font_normal
         ws['E5'].border = borda
     
         ws['B6'] = 'Período:'
         ws['B6'].font = font_titulo
         ws['B6'].alignment = align_center
         ws['B6'].fill = PatternFill(fill_type='solid', start_color='1f497d')
         ws['C6'] = periodo
         ws['C6'].font = font_normal

         

         for linha in range(5, 7):
             for col in ["B", "C", "D", "E"]:
                 ws[f"{col}{linha}"].border = borda

         # Título da seção
         ws['B9'] = titulo_secao
         ws['B9'].font = Font(name='Arial', bold=True, size=11)

         linha = 12

        
         
         for nome_grupo, lista in grupos_dict.items():
             
             
             
             celula = posicaoIndefinido.get(nome_grupo)

             # 🔹 Título do grupo
             if not "AUTOMATICA" in nome_grupo:
                ws.merge_cells(f'B{linha}:E{linha}')
             else:
                ws.merge_cells(f'B{linha}:D{linha}')
                ws[f'E{linha}'].border = Border(left=None, right=None, top=None, bottom=None)
             ws[f'B{linha}'] = (
              f'="GRUPO: "&Resumo!{celula}'
              if celula
              else f'GRUPO: {nome_grupo}'
            )
             ws[f'B{linha}'].font = Font(name='Arial', bold=True, size=10)
             ws[f'B{linha}'].alignment = align_left
            
             linha += 1

             # 🔹 Cabeçalho tabela
             headers = ["Data", "Lançamento", "Valor (R$)", "Saldo (R$)"]
             if nome_aba == "Aplicacoes":
                headers = ["Data", "Lançamento", "Aplicação (R$)", "Resgate (R$)"]
             colunas = ["B", "C", "D", "E"]

             if "AUTOMATICA" in nome_grupo:
                headers = ["Data", "Lançamento", "Rendimento"]
                colunas = ["B", "C", "D"]
                 
                 

             


             for col, texto in zip(colunas, headers):
                 cell = ws[f"{col}{linha}"]
                 cell.value = texto
                 cell.font = font_cabecalho_tabela
                 cell.alignment = align_center
                 cell.fill = fill_cabecalho
                 cell.border = border_thin

             linha += 1
             linha_inicio_dados = linha

             if not lista:
                 ws[f"B{linha}"] = "Sem registros"
                 ws[f"B{linha}"].font = Font(name='Arial', italic=True, size=10)
                 linha += 2
                

             for reg in lista:

                 valor = float(reg["valor"])

                 ws[f"B{linha}"] = reg["data"]
                 ws[f"B{linha}"].font = font_normal
                 ws[f"B{linha}"].alignment = align_center
                 ws[f"B{linha}"].border = border_thin

                 ws[f"C{linha}"] = reg["lancamento"]
                 ws[f"C{linha}"].font = font_normal
                 ws[f"C{linha}"].alignment = align_left
                 ws[f"C{linha}"].border = border_thin
                 

                 if valor < 0 and not "AUTOMATICA" in nome_grupo :
                     ws[f"D{linha}"] = valor
                     
                 ws[f"D{linha}"].font = font_negativo 
                 ws[f"D{linha}"].alignment = align_right
                 ws[f"D{linha}"].number_format = '#,##0.00'
                 ws[f"D{linha}"].border = border_thin

                 if valor > 0 and not "AUTOMATICA" in nome_grupo:
                     ws[f"E{linha}"] = valor
                 ws[f"E{linha}"].font = font_normal
                 ws[f"E{linha}"].alignment = align_right
                 ws[f"E{linha}"].number_format = '#,##0.00'
                 ws[f"E{linha}"].border = border_thin

                 if "AUTOMATICA" in nome_grupo:
                    ws[f"D{linha}"] = valor
                    ws[f"D{linha}"].font = font_normal
                    ws[f"D{linha}"].alignment = align_right
                    ws[f"D{linha}"].number_format = '#,##0.00'
                    ws[f"D{linha}"].border = border_thin
                    ws[f"E{linha}"].border = borda_esquerda

                 linha += 1

             # 🔹 TOTAL DO GRUPO
             ws.merge_cells(f'B{linha}:C{linha}')
             ws[f"B{linha}"] = "TOTAL"
             ws[f"B{linha}"].font = font_total
             ws[f"B{linha}"].alignment = align_center
             ws[f"B{linha}"].fill = fill_total
             ws[f"B{linha}"].border = border_thin
             ws[f"C{linha}"].border = border_thin

             ws[f"D{linha}"] = f'=SUM(D{linha_inicio_dados}:D{linha-1})'
             ws[f"D{linha}"].font = font_total
             ws[f"D{linha}"].alignment = align_right
             ws[f"D{linha}"].number_format = '#,##0.00'
             ws[f"D{linha}"].fill = fill_total
             ws[f"D{linha}"].border = border_thin
             aplicacaoTotal = "D" + str(linha)

             if not "AUTOMATICA" in nome_grupo:
                ws[f"E{linha}"] = f'=SUM(E{linha_inicio_dados}:E{linha-1})'
                ws[f"E{linha}"].font = font_total
                ws[f"E{linha}"].alignment = align_right
                ws[f"E{linha}"].number_format = '#,##0.00'
                ws[f"E{linha}"].fill = fill_total
                ws[f"E{linha}"].border = border_thin
                resgateTotal = "E" + str(linha)

             # ======================================================
             # SALDOS (APENAS PARA ABA APLICACOES)
             # ======================================================
             if nome_aba == "Aplicacoes" and mapa_saldos_resumo and not ("AUTOMATICA" in nome_grupo):
                 
                 
                 print(f"Calculando saldos para o grupo '{nome_grupo}'...")

                 saldoInicial = mapa_saldos_resumo[nome_grupo]["saldo_inicial"]
                 saldoFinal = mapa_saldos_resumo[nome_grupo]["saldo_final"]

                 linha += 1

                 # Saldo Inicial
                 ws[f'D{linha}'] = 'Saldo Inicial:'
                 ws[f'D{linha}'].font = font_normal
                 ws[f"D{linha}"].border = border_thin

                 ws[f'E{linha}'] = f"=Resumo!{saldoInicial}"
                 ws[f'E{linha}'].number_format = '#,##0.00'
                 ws[f"E{linha}"].border = border_thin

                 linha += 1

                 # Saldo Final
                 ws[f'D{linha}'] = 'Saldo Final:'
                 ws[f'D{linha}'].font = font_normal
                 ws[f"D{linha}"].border = border_thin

                 ws[f'E{linha}'] = f"=Resumo!{saldoFinal}"
                 ws[f'E{linha}'].number_format = '#,##0.00'
                 ws[f"E{linha}"].border = border_thin

                 linha += 1

                 # Rendimento
                 ws[f'D{linha}'] = 'Rendimento:'
                 ws[f'D{linha}'].font = font_normal
                 ws[f"D{linha}"].border = border_thin

                 ws[f'E{linha}'] = (
                     f"=Resumo!{saldoFinal}"
                     f"-(Resumo!{saldoInicial}"
                     f"+(-{aplicacaoTotal})-{resgateTotal})"
                 )
                 ws[f'E{linha}'].number_format = '#,##0.00'
                 ws[f"E{linha}"].border = border_thin

                 linha += 2 
             if "AUTOMATICA" in nome_grupo:
                linha += 2
             else:
               ws[f'E{linha}'].font = font_normal
               ws[f'E{linha}'].number_format = '#,##0.00'
               ws[f"E{linha}"].border = border_thin

         ws.column_dimensions["B"].width = 12
         ws.column_dimensions["C"].width = 50
         ws.column_dimensions["D"].width = 18
         ws.column_dimensions["E"].width = 18

         configurar_impressao(ws, linha)
     
    # ========== ABA RESUMO ==========
    if banco == "Itau":
        aplicacoes_por_grupo, lancamentos_gerais = separar_por_grupo(obj, "Itau")
        if "CDB" not in aplicacoes_por_grupo:
            aplicacoes_por_grupo["CDB"] = []
        aplicacoes_por_grupo["Indefinido"] = []
        aplicacoes_por_grupo["Indefinido2"] = []
    elif banco == "Santander":
        aplicacoes_por_grupo, lancamentos_gerais = separar_por_grupo(obj, "Santander")
        aplicacoes_por_grupo["Indefinido"] = []
        aplicacoes_por_grupo["Indefinido2"] = []
    else:
        aplicacoes_por_grupo, lancamentos_gerais = separar_por_grupo(obj, "Bradesco")
        aplicacoes_por_grupo["Indefinido"] = []
        aplicacoes_por_grupo["Indefinido2"] = []

    ws = wb.active
    ws.title = "Resumo"

    ws.sheet_view.showGridLines = False
    mapa_saldos_resumo = {}

    
    ws.merge_cells('B2:C2')
    if banco == "Itau":
        img = Image(resource_path("img/itau_header.png"))
    elif banco == "Santander":
        img = Image(resource_path("img/santander_header.png"))
    else:
        img = Image(resource_path("./img/logo_bradesco.png"))
    ws.add_image(img, 'B1')
    ws.row_dimensions[2].height = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 90

    
    ws['B4'] = 'Nome:'
    ws['B4'].font = font_titulo
    ws['B4'].alignment = align_center
    ws['B4'].fill = PatternFill(fill_type='solid', start_color='1f497d')
    ws['C4'] = nome
    ws['C4'].font = font_normal
    
    ws['B5'] = 'Agência/Conta:'
    ws['B5'].font = font_titulo
    ws['B5'].alignment = align_center
    ws['B5'].fill = PatternFill(fill_type='solid', start_color='1f497d')
    ws['C5'] = agencia_conta
    ws['C5'].font = font_normal
    
    ws['B6'] = 'Período:'
    ws['B6'].font = font_titulo
    ws['B6'].alignment = align_center
    ws['B6'].fill = PatternFill(fill_type='solid', start_color='1f497d')
    ws['C6'] = periodo
    ws['C6'].font = font_normal

    borda_simples = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
    )

    for linha in range(4, 7):          # 4 até 6
        for col in ["B", "C"]:         # colunas B e C
            ws[f"{col}{linha}"].border = borda_simples
    
    ws['B8'] = 'SALDOS'
    ws['B8'].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws['B8'].alignment = align_center
    ws['B8'].fill = PatternFill(fill_type='solid', start_color='1f497d')
    
    ws['B9'] = 'Saldo Inicial:'
    ws['B9'].font = font_normal
    ws['C9'] = saldo_inicial
    ws['C9'].font = font_normal
    ws['C9'].number_format = '#,##0.00'
    
    ws['B10'] = 'Saldo Final:'
    ws['B10'].font = font_normal
    ws['C10'] = saldo_final
    ws['C10'].font = font_normal
    ws['C10'].number_format = '#,##0.00'
    
    ws['B12'] = 'TOTALIZADORES'
    ws['B12'].font = Font(name='Arial', bold=True , size=11, color='FFFFFF')
    ws['B12'].alignment = align_center
    ws['B12'].fill = PatternFill(fill_type='solid', start_color='1f497d')
    
    ws['B13'] = 'Total Tarifas/Custas:'
    ws['B13'].font = font_normal
    ws['C13'] = sum(float(x["valor"]) for x in tarifas) if tarifas else 0
    ws['C13'].font = Font(name='Arial', size=10, color="FF0000")
    ws['C13'].number_format = '#,##0.00'
    
    ws['B14'] = 'Total Créditos:'
    ws['B14'].font = font_normal
    ws['C14'] = sum(float(x["valor"]) for x in creditos) if creditos else 0
    ws['C14'].font = Font(name='Arial', size=10, color="008000")
    ws['C14'].number_format = '#,##0.00'
    
    ws['B15'] = 'Total Débitos:'
    ws['B15'].font = font_normal
    ws['C15'] = sum(float(x["valor"]) for x in debitos) if debitos else 0
    ws['C15'].font = Font(name='Arial', size=10, color="FF0000")
    ws['C15'].number_format = '#,##0.00'
    
    ws['B16'] = 'Total Aplicações:'
    ws['B16'].font = font_normal
    ws['C16'] = sum(float(x["valor"]) for x in aplicacoes) if aplicacoes else 0
    ws['C16'].font = font_normal
    ws['C16'].number_format = '#,##0.00'

    ws["B8"].alignment = Alignment(horizontal="center")
    ws["B12"].alignment = Alignment(horizontal="center")

    # Mesclar títulos
    ws.merge_cells("B8:C8")
    ws.merge_cells("B12:C12")

    # Aplicar borda no quadro todo
    for linha in range(8, 11):
        for col in ["B", "C"]:
            ws[f"{col}{linha}"].border = borda
    for linha in range(12, 17):
        for col in ["B", "C"]:
            ws[f"{col}{linha}"].border = borda
    
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    

    linha_resumo = 18  # começa depois da parte que já existe

    ws[f"B{linha_resumo}"] = "APLICAÇÕES"
    ws[f"B{linha_resumo}"].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ws[f"B{linha_resumo}"].fill = PatternFill(fill_type='solid', start_color='1f497d')
    ws[f"B{linha_resumo}"].border = border_thin
    ws[f"B{linha_resumo}"].alignment = align_center
    ws.merge_cells(f"B{linha_resumo}:C{linha_resumo}")

    linha_resumo += 2
    posicaoIndefinido = {}


    for nome_grupo in aplicacoes_por_grupo.keys():

        if "AUTOMATICA" in nome_grupo:
            continue

        if "Indefinido" in nome_grupo:
            posicaoIndefinido[nome_grupo] = f"B{linha_resumo}"
        if "Indefinido2" in nome_grupo:
            posicaoIndefinido[nome_grupo] = f"B{linha_resumo}"
        

        # 🔹 Nome da aplicação
        ws[f"B{linha_resumo}"] = nome_grupo
        ws[f"B{linha_resumo}"].font = Font(name='Arial', bold=True, size=10)
        ws.merge_cells(f"B{linha_resumo}:C{linha_resumo}")
        ws[f"B{linha_resumo}"].border = border_thin
        ws[f"C{linha_resumo}"].border = border_thin
        

        linha_resumo += 1

        # 🔹 Saldo Inicial
        ws[f"B{linha_resumo}"] = "Saldo Inicial:"
        ws[f"C{linha_resumo}"] = 0
        ws[f"C{linha_resumo}"].number_format = '#,##0.00'

        ws[f"B{linha_resumo}"].border = border_thin
        ws[f"C{linha_resumo}"].border = border_thin

        saldo_inicial_cell = f"C{linha_resumo}"
        linha_resumo += 1

        # 🔹 Saldo Final
        ws[f"B{linha_resumo}"] = "Saldo Final:"
        ws[f"C{linha_resumo}"] = 0
        ws[f"C{linha_resumo}"].number_format = '#,##0.00'

        ws[f"B{linha_resumo}"].border = border_thin
        ws[f"C{linha_resumo}"].border = border_thin

        saldo_final_cell = f"C{linha_resumo}"

        # Armazenar os valores das células de saldo inicial e final no mapa
        mapa_saldos_resumo[nome_grupo] = {
            "saldo_inicial": saldo_inicial_cell,
            "saldo_final": saldo_final_cell
        }

        linha_resumo += 2    

    configurar_impressao(ws, 18)

    


    # ========== CRIAR ABAS ==========
    montar_aba("Tarifas", tarifas, "APURAÇÃO TARIFAS BANCÁRIAS")
    montar_aba_grupos("Aplicacoes",aplicacoes_por_grupo , "APLICAÇÕES", mapa_saldos_resumo)
    montar_aba("Geral", lancamentos_gerais, "GERAL")
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    wb.save(caminho)
    print(f"Arquivo {banco} estilizado gerado: {caminho}")

